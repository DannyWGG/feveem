from ninja import Router
from datetime import datetime

from openpyxl import Workbook
from openpyxl.drawing.image import Image
from openpyxl.chart import PieChart, Reference
from openpyxl.styles import Alignment

from io import BytesIO
from django.http import HttpResponse
from django.db.models import Count

from apps.feveem.models.asistente import Asistente

router = Router()

@router.get("/general", response=None)
def reporte_general(request):

 # Crear un nuevo archivo Excel
    wb = Workbook()
    hoja = wb.active
    hoja.title = "Reporte General"

    # Configurar encabezado con imagen
    try:
        logo = Image("static/img/barra.png")  # Ajusta la ruta según tu proyecto
        logo.width, logo.height = 2220, 90
        hoja.add_image(logo, "A1")
    except FileNotFoundError:
        pass  # Si la imagen no existe, continuar sin ella
    
    hoja['B6']= "Voceros"
    hoja['E6']= "Extracurriculares"

    hoja['E6'].alignment = Alignment(horizontal='center', vertical='center')
    hoja['B6'].alignment = Alignment(horizontal='center', vertical='center')

    hoja.merge_cells('B6:D6')
    hoja.merge_cells('E6:G6')

    # Encabezados
    encabezados = [
        "Estado", "Contralor", "Integrador", "Activista", "Deportes", "Cultural", "Científico"]
    
    fila_inicial = 7
    for col_index, encabezado in enumerate(encabezados, start=1):  # Comienza en columna A (1)
        hoja.cell(row=fila_inicial, column=col_index, value=encabezado)

    estados = [
        "AMAZONAS", "ANZOATEGUI", "APURE", "ARAGUA", "BARINAS", "BOLIVAR", "CARABOBO",
        "COJEDES", "DELTA AMACURO", "DISTRITO CAPITAL", "FALCON", "GUÁRICO", "LARA",
        "MERIDA", "MIRANDA", "MONAGAS", "NUEVA ESPARTA", "PORTUGUESA", "SUCRE",
        "TÁCHIRA", "TRUJILLO", "LA GUAIRA", "YARACUY", "ZULIA"
    ]

    # Contar asistentes con vocería de tipo "contralor" por estado
    conteo_contralor = (
        Asistente.objects.filter(voceria__descripcion="CONTROLADOR")
        .values("estado")
        .annotate(total=Count("id"))
    )

     # Crear un diccionario para mapear los conteos por estado
    conteo_por_estado = {estado["estado"]: estado["total"] for estado in conteo_contralor}

    # Contar asistentes con vocería de tipo "integrador" por estado
    conteo_integrador = (
        Asistente.objects.filter(voceria__descripcion="INTEGRADOR")
        .values("estado")
        .annotate(total=Count("id"))
    )

     # Crear un diccionario para mapear los conteos por estado
    conteo_por_estado_int = {estado["estado"]: estado["total"] for estado in conteo_integrador}

    # Contar asistentes con vocería de tipo "activista" por estado
    conteo_activista = (
        Asistente.objects.filter(voceria__descripcion="ACTIVISTA")
        .values("estado")
        .annotate(total=Count("id"))
    )

     # Crear un diccionario para mapear los conteos por estado
    conteo_por_estado_act = {estado["estado"]: estado["total"] for estado in conteo_activista}

    # Contar asistentes con actividaes extracurricular de tipo "deportes" por estado
    conteo_deportes = (
        Asistente.objects.filter(extra_curricular__descripcion="DEPORTES")
        .values("estado")
        .annotate(total=Count("id"))
    )

     # Crear un diccionario para mapear los conteos por estado
    conteo_por_estado_dep = {estado["estado"]: estado["total"] for estado in conteo_deportes}

    # Contar asistentes con actividaes extracurricular de tipo "cultural" por estado
    conteo_cultural = (
        Asistente.objects.filter(extra_curricular__descripcion="CULTURAL")
        .values("estado")
        .annotate(total=Count("id"))
    )

     # Crear un diccionario para mapear los conteos por estado
    conteo_por_estado_cul = {estado["estado"]: estado["total"] for estado in conteo_cultural}

    # Contar asistentes con actividaes extracurricular de tipo "Cientifico" por estado
    conteo_cientifico = (
        Asistente.objects.filter(extra_curricular__descripcion="CIENTÍFICO")
        .values("estado")
        .annotate(total=Count("id"))
    )

     # Crear un diccionario para mapear los conteos por estado
    conteo_por_estado_cien = {estado["estado"]: estado["total"] for estado in conteo_cientifico}


    # Insertar los estados y conteos en la hoja de Excel
    for i, estado in enumerate(estados):
        hoja.cell(row=i + 8, column=1, value=estado)  # Nombre del estado
        hoja.cell(row=i + 8, column=2, value=conteo_por_estado.get(estado, 0)) # Conteo de "contralor"
        hoja.cell(row=i + 8, column=3, value=conteo_por_estado_int.get(estado, 0))  # Conteo de "integrador"
        hoja.cell(row=i + 8, column=4, value=conteo_por_estado_act.get(estado, 0))  # Conteo de "activista"
        hoja.cell(row=i + 8, column=5, value=conteo_por_estado_dep.get(estado, 0))  # Conteo de "deporte"
        hoja.cell(row=i + 8, column=6, value=conteo_por_estado_cul.get(estado, 0))  # Conteo de "activista"
        hoja.cell(row=i + 8, column=7, value=conteo_por_estado_cien.get(estado, 0))  # Conteo de "activista"

    # Ajustar el ancho de columnas para que el contenido sea visible
    for col in hoja.columns:
        col_letter = col[0].column_letter  # Letra de la columna
        hoja.column_dimensions[col_letter].width = 15  # Ancho fijo


    # Crear una nueva hoja para gráficos
    hoja_graficos = wb.create_sheet(title="Gráficos")

    # Insertar gráfico por vocerías
    for idx, (voceria, col) in enumerate(zip(
        ["Contralor", "Integrador", "Activista", "Deportes", "Cultural", "Científico"], 
        range(2, 8)
    )):
        # Crear el gráfico
        chart = PieChart()
        chart.title = f"Vocería: {voceria}"
        
        # Definir los datos
        datos = Reference(hoja, min_col=col, min_row=7, max_row=7 + len(estados))
        etiquetas = Reference(hoja, min_col=1, min_row=8, max_row=8 + len(estados) - 1)
        
        chart.add_data(datos, titles_from_data=False)
        chart.set_categories(etiquetas)
        
        # Colocar el gráfico en la hoja
        hoja_graficos.add_chart(chart, f"A{idx * 15 + 1}")
    
    # Guardar el archivo en memoria
    """ buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0) """
    
    # Instancias para la generacion del reporte por el metodo HTTP
    ctime       = datetime.now().strftime('%d-%m-%Y')
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=Reporte-General - % s.xlsx' % ctime
    wb.save(response)

    return response
